from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy.exc import IntegrityError
from sqlalchemy.orm import Session

from app.auth import get_current_user
from app.database import get_db
from app.models import Equipment, User
from app.schemas import (
    EquipmentCreate,
    EquipmentImportResult,
    EquipmentListResponse,
    EquipmentResponse,
    EquipmentUpdate,
)

router = APIRouter(prefix="/api/equipment", tags=["equipment"])

VALID_STATUSES = {"运行", "停机", "维修", "报废"}
# 导入文件大小上限，防止超大 Excel 占满内存
MAX_IMPORT_BYTES = 10 * 1024 * 1024
IMPORT_READ_CHUNK = 1024 * 1024

EXPORT_HEADERS = [
    "设备编号",
    "设备名称",
    "规格型号",
    "使用部门",
    "安装位置",
    "设备状态",
    "购置日期",
    "启用日期",
    "供应商/制造商",
    "备注",
]


async def _read_upload_limited(file: UploadFile, max_bytes: int = MAX_IMPORT_BYTES) -> bytes:
    """分块读取上传文件，超过上限则拒绝，避免整文件读入导致 OOM。"""
    declared = getattr(file, "size", None)
    if declared is not None and declared > max_bytes:
        raise HTTPException(
            status_code=400,
            detail=f"文件过大，请上传不超过 {max_bytes // (1024 * 1024)}MB 的 Excel",
        )

    chunks: list[bytes] = []
    total = 0
    while True:
        chunk = await file.read(IMPORT_READ_CHUNK)
        if not chunk:
            break
        total += len(chunk)
        if total > max_bytes:
            raise HTTPException(
                status_code=400,
                detail=f"文件过大，请上传不超过 {max_bytes // (1024 * 1024)}MB 的 Excel",
            )
        chunks.append(chunk)
    return b"".join(chunks)


def _get_equipment_or_404(equipment_id: int, db: Session) -> Equipment:
    equipment = db.query(Equipment).filter(Equipment.id == equipment_id).first()
    if not equipment:
        raise HTTPException(status_code=404, detail="设备不存在")
    return equipment


def _apply_filters(
    query,
    equipment_code: str | None,
    name: str | None,
    department: str | None,
    status: str | None,
    search: str | None,
):
    if equipment_code:
        query = query.filter(Equipment.equipment_code.ilike(f"%{equipment_code}%"))
    if name:
        query = query.filter(Equipment.name.ilike(f"%{name}%"))
    if department:
        query = query.filter(Equipment.department.ilike(f"%{department}%"))
    if status:
        query = query.filter(Equipment.status == status)
    if search:
        pattern = f"%{search}%"
        query = query.filter(
            (Equipment.equipment_code.ilike(pattern))
            | (Equipment.name.ilike(pattern))
            | (Equipment.department.ilike(pattern))
        )
    return query


def _parse_date(value) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, date):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


@router.get("", response_model=EquipmentListResponse)
def list_equipment(
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=1000),
    equipment_code: str | None = Query(None),
    name: str | None = Query(None),
    department: str | None = Query(None),
    status: str | None = Query(None),
    search: str | None = Query(None),
    _current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = db.query(Equipment)
    query = _apply_filters(query, equipment_code, name, department, status, search)
    query = query.order_by(Equipment.created_at.desc())
    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()
    return EquipmentListResponse(
        items=items,
        total=total,
        page=page,
        page_size=page_size,
    )


@router.get("/export")
def export_equipment(
    equipment_code: str | None = Query(None),
    name: str | None = Query(None),
    department: str | None = Query(None),
    status: str | None = Query(None),
    search: str | None = Query(None),
    _current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    query = db.query(Equipment)
    query = _apply_filters(query, equipment_code, name, department, status, search)
    items = query.order_by(Equipment.equipment_code).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "设备台账"
    ws.append(EXPORT_HEADERS)
    for item in items:
        ws.append([
            item.equipment_code,
            item.name,
            item.spec_model or "",
            item.department or "",
            item.location or "",
            item.status,
            item.purchase_date.isoformat() if item.purchase_date else "",
            item.commission_date.isoformat() if item.commission_date else "",
            item.supplier or "",
            item.remark or "",
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"equipment_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.get("/{equipment_id}", response_model=EquipmentResponse)
def get_equipment(
    equipment_id: int,
    _current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    return _get_equipment_or_404(equipment_id, db)


@router.post("", response_model=EquipmentResponse, status_code=201)
def create_equipment(
    payload: EquipmentCreate,
    _current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    existing = (
        db.query(Equipment)
        .filter(Equipment.equipment_code == payload.equipment_code)
        .first()
    )
    if existing:
        raise HTTPException(status_code=409, detail="设备编号已存在")

    equipment = Equipment(**payload.model_dump())
    db.add(equipment)
    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="设备编号已存在") from None
    db.refresh(equipment)
    return equipment


@router.put("/{equipment_id}", response_model=EquipmentResponse)
def update_equipment(
    equipment_id: int,
    payload: EquipmentUpdate,
    _current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    equipment = _get_equipment_or_404(equipment_id, db)
    update_data = payload.model_dump(exclude_unset=True)

    if (
        "equipment_code" in update_data
        and update_data["equipment_code"] != equipment.equipment_code
    ):
        duplicate = (
            db.query(Equipment)
            .filter(
                Equipment.equipment_code == update_data["equipment_code"],
                Equipment.id != equipment_id,
            )
            .first()
        )
        if duplicate:
            raise HTTPException(status_code=409, detail="设备编号已存在")

    for field, value in update_data.items():
        setattr(equipment, field, value)
    equipment.updated_at = datetime.utcnow()

    try:
        db.commit()
    except IntegrityError:
        db.rollback()
        raise HTTPException(status_code=409, detail="设备编号已存在") from None
    db.refresh(equipment)
    return equipment


@router.delete("/{equipment_id}", status_code=204)
def delete_equipment(
    equipment_id: int,
    _current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    equipment = _get_equipment_or_404(equipment_id, db)
    db.delete(equipment)
    db.commit()


@router.post("/import", response_model=EquipmentImportResult)
async def import_equipment(
    file: UploadFile = File(...),
    _current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="请上传 Excel 文件（.xlsx）")

    content = await _read_upload_limited(file)
    if not content:
        raise HTTPException(status_code=400, detail="上传文件为空")

    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(status_code=400, detail="Excel 文件无数据行")

    created = 0
    skipped = 0
    errors: list[str] = []

    for idx, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        code = str(row[0]).strip() if row[0] is not None else ""
        name_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""

        if not code or not name_val:
            errors.append(f"第 {idx} 行：设备编号和名称不能为空")
            continue

        existing = db.query(Equipment).filter(Equipment.equipment_code == code).first()
        if existing:
            skipped += 1
            continue

        status_val = str(row[5]).strip() if len(row) > 5 and row[5] is not None else "运行"
        if status_val not in VALID_STATUSES:
            status_val = "运行"

        equipment = Equipment(
            equipment_code=code,
            name=name_val,
            spec_model=str(row[2]).strip() if len(row) > 2 and row[2] is not None else None,
            department=str(row[3]).strip() if len(row) > 3 and row[3] is not None else None,
            location=str(row[4]).strip() if len(row) > 4 and row[4] is not None else None,
            status=status_val,
            purchase_date=_parse_date(row[6]) if len(row) > 6 else None,
            commission_date=_parse_date(row[7]) if len(row) > 7 else None,
            supplier=str(row[8]).strip() if len(row) > 8 and row[8] is not None else None,
            remark=str(row[9]).strip() if len(row) > 9 and row[9] is not None else None,
        )
        # 使用 savepoint：单行冲突只回滚本行，不影响已成功 flush 的其他行
        try:
            with db.begin_nested():
                db.add(equipment)
                db.flush()
            created += 1
        except IntegrityError:
            skipped += 1
            errors.append(f"第 {idx} 行：设备编号 {code} 已存在")

    db.commit()
    return EquipmentImportResult(created=created, skipped=skipped, errors=errors)
